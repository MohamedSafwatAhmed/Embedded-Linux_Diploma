from openpyxl import Workbook
wb=Workbook()
ws=wb.active
f=open('Dio.h','r')
i=1
for file in f:
    ws[f'A{i}']=file.strip()
    str='IDX%03d'%i
    ws[f'B{i}']=str
    i=i+1    
f.close()

wb.save("sample.xlsx")